import logging
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    ConversationHandler,
    ContextTypes,
    filters,
)
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import os
from dotenv import load_dotenv

# 1. ØªØ­Ù…ÙŠÙ„ Ù…ØªØºÙŠØ±Ø§Øª Ø§Ù„Ø¨ÙŠØ¦Ø©
load_dotenv()

# 2. Ø¬Ù„Ø¨ Ø§Ù„ØªÙˆÙƒÙ† Ù…Ù† Ø§Ù„Ù†Ø¸Ø§Ù…
BOT_TOKEN = os.getenv("BOT_TOKEN")

# ØªÙØ¹ÙŠÙ„ Ø§Ù„Ù„ÙˆØ¬ÙŠÙ†Ø¬
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# --- Ø§Ù„Ø«ÙˆØ§Ø¨Øª ÙˆØ§Ù„Ù…Ø¬Ù„Ø¯Ø§Øª ---
MAIN_MENU, VISIT_TYPE, DOCTOR_NAME, LOCATION, SPECIALTY, PRODUCTS, COMMENT = range(7)
PHARMACY_NAME, PHARMACY_ADDRESS, PHARMACY_PRODUCTS, PHARMACY_COMMENT = range(7, 11)
FIRST_NAME_INPUT, LAST_NAME_INPUT = 11, 12

REPORTS_DIR = "reports"
if not os.path.exists(REPORTS_DIR):
    os.makedirs(REPORTS_DIR)

# --- ÙƒÙ„Ø§Ø³ ExcelHandler (Ù…Ø¹Ø¯Ù„) ---
class ExcelHandler:
    @staticmethod
    def get_today_filename(user_id, first_name):
        """Ø¥Ù†Ø´Ø§Ø¡ Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù Ø¨Ù†Ø§Ø¡Ù‹ Ø¹Ù„Ù‰ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø£ÙˆÙ„ Ùˆ user_id"""
        today = datetime.now()
        day_name = today.strftime("%a")
        date_str = today.strftime("%d-%b")
        username = f"{first_name}{user_id}"
        return f"{username}_Report_{day_name}_{date_str}.xlsx"
    
    @staticmethod
    def create_new_report(user_id, first_name, full_name):
        """Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯ Ù…Ø¹ Ø§Ù„Ø§Ø³Ù… Ø§Ù„ÙƒØ§Ù…Ù„ ÙÙŠ Ø§Ù„Ø¯Ø§Ø®Ù„"""
        filename = ExcelHandler.get_today_filename(user_id, first_name)
        filepath = os.path.join(REPORTS_DIR, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Report"
        
        # Ø§Ù„ØªÙ†Ø³ÙŠÙ‚Ø§Øª
        header_fill = PatternFill("solid", fgColor="FFFF00")
        blue_fill = PatternFill("solid", fgColor="31859B")
        orange_fill = PatternFill("solid", fgColor="FABF8F")
        section_fill = PatternFill("solid", fgColor="C6E0B4")
        center = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        bold = Font(bold=True)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Ø§Ù„Ø¹Ù†ÙˆØ§Ù† Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠ
        ws.merge_cells("A2:F2")
        ws["A2"].value = "Daily Report"
        ws["A2"].font = Font(bold=True, size=14)
        ws["A2"].alignment = center
        ws["A2"].fill = header_fill
        
        # Ø§Ù„Ø§Ø³Ù… ÙˆØ§Ù„ØªØ§Ø±ÙŠØ®
        ws["A4"].value = "Name:"
        ws["A5"].value = "Date:"
        ws["A4"].font = ws["A5"].font = bold
        
        ws.merge_cells("B4:F4")
        ws.merge_cells("B5:F5")
        ws["B4"].value = full_name  # Ø§Ù„Ø§Ø³Ù… Ø§Ù„ÙƒØ§Ù…Ù„ Ù‡Ù†Ø§
        ws["B5"].value = datetime.now().strftime("%d/%m/%Y")
        ws["B5"].alignment = left_align
        
        for col in ["A", "B"]:
            ws[f"{col}4"].fill = blue_fill
            ws[f"{col}5"].fill = orange_fill
        
        # Headers Ù„Ù„Ø²ÙŠØ§Ø±Ø§Øª
        headers = ["A.M / P.M", "Doctor Name", "Hospital", "Specialist", "Product", "Comment"]
        header_row = 7
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        
        # Ù‚Ø³Ù… A.M
        ws.merge_cells("A8:A14")
        ws["A8"].value = "A.M"
        ws["A8"].alignment = center
        ws["A8"].font = bold
        ws["A8"].fill = section_fill
        
        for r in range(8, 15):
            for c in range(2, 7):
                ws.cell(row=r, column=c).border = border
        
        for c in range(1, 7):
            ws.cell(row=15, column=c).fill = orange_fill
        
        # Ù‚Ø³Ù… P.M
        ws.merge_cells("A16:A28")
        ws["A16"].value = "P.M"
        ws["A16"].alignment = center
        ws["A16"].font = bold
        ws["A16"].fill = section_fill
        
        for r in range(16, 29):
            for c in range(2, 7):
                ws.cell(row=r, column=c).border = border
        
        for c in range(1, 7):
            ws.cell(row=29, column=c).fill = orange_fill
        
        # Ù‚Ø³Ù… PHARMACY
        ws.merge_cells("A30:A37")
        ws["A30"].value = "PHARMACY"
        ws["A30"].alignment = center
        ws["A30"].font = bold
        ws["A30"].fill = header_fill
        
        ph_headers = ["Pharmacy Name", "Address", "Products", "Comments"]
        ph_cols = [2, 3, 4, 6]
        for col, h in zip(ph_cols, ph_headers):
            cell = ws.cell(row=30, column=col, value=h)
            cell.font = bold
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center
        
        for r in range(30, 38):
            ws.merge_cells(f"D{r}:E{r}")
        
        for r in range(31, 38):
            for c in range(2, 7):
                ws.cell(row=r, column=c).border = border
        
        # Ø¹Ø±Ø¶ Ø§Ù„Ø£Ø¹Ù…Ø¯Ø©
        widths = [15, 25, 20, 20, 20, 30]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        
        wb.save(filepath)
        return filepath, True
    
    @staticmethod
    def add_visit(user_id, first_name, visit_type, data):
        """Ø¥Ø¶Ø§ÙØ© Ø²ÙŠØ§Ø±Ø© Ù„Ù„ØªÙ‚Ø±ÙŠØ±"""
        filename = ExcelHandler.get_today_filename(user_id, first_name)
        filepath = os.path.join(REPORTS_DIR, filename)
        
        if not os.path.exists(filepath):
            return None  # Ø§Ù„Ù…Ù„Ù ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯
        
        wb = load_workbook(filepath)
        ws = wb.active
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        if visit_type == "AM":
            for row in range(8, 15):
                if not ws.cell(row=row, column=2).value:
                    ws.cell(row=row, column=2).value = data.get("Dr", "")
                    ws.cell(row=row, column=3).value = data.get("Hospital", "")
                    ws.cell(row=row, column=4).value = data.get("Specialty", "")
                    ws.cell(row=row, column=5).value = data.get("Products", "")
                    ws.cell(row=row, column=6).value = data.get("Comment", "")
                    for c in range(2, 7):
                        ws.cell(row=row, column=c).border = border
                    break
        
        elif visit_type == "PM":
            for row in range(16, 29):
                if not ws.cell(row=row, column=2).value:
                    ws.cell(row=row, column=2).value = data.get("Dr", "")
                    ws.cell(row=row, column=3).value = data.get("Area", "")
                    ws.cell(row=row, column=4).value = data.get("Specialty", "")
                    ws.cell(row=row, column=5).value = data.get("Products", "")
                    ws.cell(row=row, column=6).value = data.get("Comment", "")
                    for c in range(2, 7):
                        ws.cell(row=row, column=c).border = border
                    break
        
        elif visit_type == "PHARMACY":
            for row in range(31, 38):
                if not ws.cell(row=row, column=2).value:
                    ws.cell(row=row, column=2).value = data.get("Pharmacy", "")
                    ws.cell(row=row, column=3).value = data.get("Address", "")
                    ws.merge_cells(f"D{row}:E{row}")
                    ws.cell(row=row, column=4).value = data.get("Products", "")
                    ws.cell(row=row, column=6).value = data.get("Comment", "")
                    for c in range(2, 7):
                        ws.cell(row=row, column=c).border = border
                    break
        
        wb.save(filepath)
        return filepath

# --- ÙˆØ¸Ø§Ø¦Ù Ø§Ù„Ø¨ÙˆØª ---
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["ğŸ“Š Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯"],
        ["âœ… ØªØ³Ø¬ÙŠÙ„ Ø²ÙŠØ§Ø±Ø© Ø¬Ø¯ÙŠØ¯Ø©"],
        ["ğŸ“¤ Ø¥Ø±Ø³Ø§Ù„ Ø§Ù„ØªÙ‚Ø±ÙŠØ±"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "ğŸ¤– *Medical Rep Bot*\n\nÙ…Ø±Ø­Ø¨Ø§Ù‹! Ø§Ø®ØªØ± Ù…Ù† Ø§Ù„Ù‚Ø§Ø¦Ù…Ø©:",
        reply_markup=reply_markup,
        parse_mode='Markdown'
    )
    return MAIN_MENU

async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text
    
    if choice == "ğŸ“Š Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯":
        await update.message.reply_text(
            "ğŸ‘¤ Ø£Ø¯Ø®Ù„ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø£ÙˆÙ„:",
            reply_markup=ReplyKeyboardRemove()
        )
        return FIRST_NAME_INPUT
    
    elif choice == "âœ… ØªØ³Ø¬ÙŠÙ„ Ø²ÙŠØ§Ø±Ø© Ø¬Ø¯ÙŠØ¯Ø©":
        keyboard = [
            ["ğŸŒ… A.M Visit"],
            ["ğŸŒ† P.M Visit"],
            ["ğŸ’Š Pharmacy Visit"],
            ["ğŸ”™ Ø±Ø¬ÙˆØ¹"]
        ]
        reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
        await update.message.reply_text("Ø§Ø®ØªØ± Ù†ÙˆØ¹ Ø§Ù„Ø²ÙŠØ§Ø±Ø©:", reply_markup=reply_markup)
        return VISIT_TYPE
    
    elif choice == "ğŸ“¤ Ø¥Ø±Ø³Ø§Ù„ Ø§Ù„ØªÙ‚Ø±ÙŠØ±":
        return await send_report(update, context)

async def first_name_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ø§Ø³ØªÙ‚Ø¨Ø§Ù„ Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø£ÙˆÙ„"""
    context.user_data['first_name'] = update.message.text.strip()
    await update.message.reply_text("ğŸ‘¤ Ø£Ø¯Ø®Ù„ Ø¨Ø§Ù‚ÙŠ Ø§Ù„Ø§Ø³Ù…:")
    return LAST_NAME_INPUT

async def last_name_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Ø§Ø³ØªÙ‚Ø¨Ø§Ù„ Ø¨Ø§Ù‚ÙŠ Ø§Ù„Ø§Ø³Ù… ÙˆØ¥Ù†Ø´Ø§Ø¡ Ø§Ù„ØªÙ‚Ø±ÙŠØ±"""
    last_name = update.message.text.strip()
    first_name = context.user_data['first_name']
    full_name = f"{first_name} {last_name}"
    user_id = update.effective_user.id
    
    # Ø­ÙØ¸ Ø§Ù„Ù…Ø¹Ù„ÙˆÙ…Ø§Øª
    context.user_data['full_name'] = full_name
    context.user_data['user_id'] = user_id
    
    # Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„ØªÙ‚Ø±ÙŠØ±
    filepath, is_new = ExcelHandler.create_new_report(user_id, first_name, full_name)
    filename = os.path.basename(filepath)
    
    if is_new:
        await update.message.reply_text(
            f"âœ… *ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„ØªÙ‚Ø±ÙŠØ± Ø¨Ù†Ø¬Ø§Ø­!*\n\n"
            f"ğŸ“„ Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù: `{filename}`\n"
            f"ğŸ‘¤ Ø§Ù„Ø§Ø³Ù…: {full_name}\n"
            f"ğŸ†” User ID: {user_id}",
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            f"â„¹ï¸ *Ø§Ù„ØªÙ‚Ø±ÙŠØ± Ù…ÙˆØ¬ÙˆØ¯ Ø¨Ø§Ù„ÙØ¹Ù„*\n\nğŸ“„ Ø§Ø³Ù… Ø§Ù„Ù…Ù„Ù: `{filename}`",
            parse_mode='Markdown'
        )
    
    return await start(update, context)

async def visit_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    choice = update.message.text
    
    if choice == "ğŸ”™ Ø±Ø¬ÙˆØ¹":
        return await start(update, context)
    
    if choice == "ğŸŒ… A.M Visit":
        context.user_data['visit_type'] = "AM"
        context.user_data['location_label'] = "Ø§Ù„Ù…Ø³ØªØ´ÙÙ‰"
    elif choice == "ğŸŒ† P.M Visit":
        context.user_data['visit_type'] = "PM"
        context.user_data['location_label'] = "Ø§Ù„Ù…Ù†Ø·Ù‚Ø©"
    elif choice == "ğŸ’Š Pharmacy Visit":
        context.user_data['visit_type'] = "PHARMACY"
        await update.message.reply_text(
            "ğŸª Ø£Ø¯Ø®Ù„ Ø§Ø³Ù… Ø§Ù„ØµÙŠØ¯Ù„ÙŠØ©:",
            reply_markup=ReplyKeyboardRemove()
        )
        return PHARMACY_NAME
    
    await update.message.reply_text(
        "ğŸ‘¤ Ø£Ø¯Ø®Ù„ Ø§Ø³Ù… Ø§Ù„Ø¯ÙƒØªÙˆØ±:",
        reply_markup=ReplyKeyboardRemove()
    )
    return DOCTOR_NAME

async def doctor_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['doctor_name'] = update.message.text
    location_label = context.user_data['location_label']
    await update.message.reply_text(f"ğŸ¥ Ø£Ø¯Ø®Ù„ {location_label}:")
    return LOCATION

async def location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['location'] = update.message.text
    await update.message.reply_text("ğŸ©º Ø£Ø¯Ø®Ù„ ØªØ®ØµØµ Ø§Ù„Ø¯ÙƒØªÙˆØ±:")
    return SPECIALTY

async def specialty(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['specialty'] = update.message.text
    await update.message.reply_text("ğŸ’Š Ø£Ø¯Ø®Ù„ Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª (Ø§ÙØµÙ„ Ø¨ÙŠÙ†Ù‡Ø§ Ø¨ÙØ§ØµÙ„Ø©):")
    return PRODUCTS

async def products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['products'] = update.message.text
    keyboard = [["â­ï¸ ØªØ®Ø·ÙŠ"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "ğŸ’¬ Ø£Ø¯Ø®Ù„ Ø§Ù„ØªØ¹Ù„ÙŠÙ‚ (Ø£Ùˆ Ø§Ø¶ØºØ· ØªØ®Ø·ÙŠ):",
        reply_markup=reply_markup
    )
    return COMMENT

async def comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    comment_text = update.message.text
    context.user_data['comment'] = comment_text if comment_text != "â­ï¸ ØªØ®Ø·ÙŠ" else ""
    
    visit_type = context.user_data['visit_type']
    location_label = "Hospital" if visit_type == "AM" else "Area"
    
    data = {
        "Dr": context.user_data['doctor_name'],
        location_label: context.user_data['location'],
        "Specialty": context.user_data['specialty'],
        "Products": context.user_data['products'],
        "Comment": context.user_data.get('comment', '')
    }
    
    # Ø§Ø³ØªØ®Ø¯Ø§Ù… user_id Ùˆ first_name Ù…Ù† Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…
    user_id = update.effective_user.id
    first_name = context.user_data.get('first_name', 'User')
    
    filepath = ExcelHandler.add_visit(user_id, first_name, visit_type, data)
    
    if filepath:
        await update.message.reply_text(
            f"âœ… *ØªÙ… ØªØ³Ø¬ÙŠÙ„ Ø§Ù„Ø²ÙŠØ§Ø±Ø© Ø¨Ù†Ø¬Ø§Ø­!*\n\n"
            f"ğŸ“„ ØªÙ… Ø§Ù„Ø­ÙØ¸ ÙÙŠ: `{os.path.basename(filepath)}`",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            "âš ï¸ *Ø®Ø·Ø£: Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø§Ù„ØªÙ‚Ø±ÙŠØ±!*\n\n"
            "Ù‚Ù… Ø¨Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯ Ø£ÙˆÙ„Ø§Ù‹.",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    
    return await start(update, context)

async def pharmacy_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['pharmacy_name'] = update.message.text
    await update.message.reply_text("ğŸ“ Ø£Ø¯Ø®Ù„ Ø¹Ù†ÙˆØ§Ù† Ø§Ù„ØµÙŠØ¯Ù„ÙŠØ©:")
    return PHARMACY_ADDRESS

async def pharmacy_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['pharmacy_address'] = update.message.text
    await update.message.reply_text("ğŸ’Š Ø£Ø¯Ø®Ù„ Ø£Ø³Ù…Ø§Ø¡ Ø§Ù„Ù…Ù†ØªØ¬Ø§Øª:")
    return PHARMACY_PRODUCTS

async def pharmacy_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['pharmacy_products'] = update.message.text
    keyboard = [["â­ï¸ ØªØ®Ø·ÙŠ"]]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text(
        "ğŸ’¬ Ø£Ø¯Ø®Ù„ Ø§Ù„ØªØ¹Ù„ÙŠÙ‚ (Ø£Ùˆ Ø§Ø¶ØºØ· ØªØ®Ø·ÙŠ):",
        reply_markup=reply_markup
    )
    return PHARMACY_COMMENT

async def pharmacy_comment(update: Update, context: ContextTypes.DEFAULT_TYPE):
    comment_text = update.message.text
    context.user_data['pharmacy_comment'] = comment_text if comment_text != "â­ï¸ ØªØ®Ø·ÙŠ" else ""
    
    data = {
        "Pharmacy": context.user_data['pharmacy_name'],
        "Address": context.user_data['pharmacy_address'],
        "Products": context.user_data['pharmacy_products'],
        "Comment": context.user_data.get('pharmacy_comment', '')
    }
    
    user_id = update.effective_user.id
    first_name = context.user_data.get('first_name', 'User')
    
    filepath = ExcelHandler.add_visit(user_id, first_name, "PHARMACY", data)
    
    if filepath:
        await update.message.reply_text(
            f"âœ… *ØªÙ… ØªØ³Ø¬ÙŠÙ„ Ø²ÙŠØ§Ø±Ø© Ø§Ù„ØµÙŠØ¯Ù„ÙŠØ© Ø¨Ù†Ø¬Ø§Ø­!*\n\n"
            f"ğŸ“„ ØªÙ… Ø§Ù„Ø­ÙØ¸ ÙÙŠ: `{os.path.basename(filepath)}`",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    else:
        await update.message.reply_text(
            "âš ï¸ *Ø®Ø·Ø£: Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ Ø§Ù„ØªÙ‚Ø±ÙŠØ±!*\n\n"
            "Ù‚Ù… Ø¨Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯ Ø£ÙˆÙ„Ø§Ù‹.",
            reply_markup=ReplyKeyboardRemove(),
            parse_mode='Markdown'
        )
    
    return await start(update, context)

async def send_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    first_name = context.user_data.get('first_name', 'User')
    
    filename = ExcelHandler.get_today_filename(user_id, first_name)
    filepath = os.path.join(REPORTS_DIR, filename)
    
    if not os.path.exists(filepath):
        await update.message.reply_text(
            "âš ï¸ *Ù„Ø§ ÙŠÙˆØ¬Ø¯ ØªÙ‚Ø±ÙŠØ± Ù„Ù„ÙŠÙˆÙ…!*\n\nÙ‚Ù… Ø¨Ø¥Ù†Ø´Ø§Ø¡ ØªÙ‚Ø±ÙŠØ± Ø¬Ø¯ÙŠØ¯ Ø£ÙˆÙ„Ø§Ù‹.",
            parse_mode='Markdown'
        )
        return await start(update, context)
    
    waiting_msg = await update.message.reply_text("â³ Ø¬Ø§Ø±ÙŠ Ø¥Ø±Ø³Ø§Ù„ Ø§Ù„ØªÙ‚Ø±ÙŠØ±...")
    
    try:
        with open(filepath, 'rb') as file:
            await update.message.reply_document(
                document=file,
                filename=filename,
                caption=f"ğŸ“Š *ØªÙ‚Ø±ÙŠØ± Ø§Ù„ÙŠÙˆÙ…*\n\nğŸ“… {datetime.now().strftime('%d %B %Y')}",
                parse_mode='Markdown'
            )
        await waiting_msg.delete()
        await update.message.reply_text("âœ… ØªÙ… Ø¥Ø±Ø³Ø§Ù„ Ø§Ù„ØªÙ‚Ø±ÙŠØ± Ø¨Ù†Ø¬Ø§Ø­!")
    except Exception as e:
        await waiting_msg.delete()
        await update.message.reply_text(f"âŒ Ø®Ø·Ø£: {str(e)}")
    
    return await start(update, context)

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "ØªÙ… Ø§Ù„Ø¥Ù„ØºØ§Ø¡. Ø§Ø³ØªØ®Ø¯Ù… /start Ù„Ù„Ø¨Ø¯Ø¡ Ù…Ø¬Ø¯Ø¯Ø§Ù‹.",
        reply_markup=ReplyKeyboardRemove()
    )
    context.user_data.clear()
    return ConversationHandler.END

# --- ÙˆØ¸ÙŠÙØ© Ø§Ù„ØªØ´ØºÙŠÙ„ Ø§Ù„Ø±Ø¦ÙŠØ³ÙŠØ© ---
def main():
    """ØªØ´ØºÙŠÙ„ Ø§Ù„Ø¨ÙˆØª Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„ØªÙˆÙƒÙ† Ù…Ù† Ù…ØªØºÙŠØ±Ø§Øª Ø§Ù„Ø¨ÙŠØ¦Ø©"""
    
    if not BOT_TOKEN:
        print("âŒ Ø®Ø·Ø£: Ù„Ù… ÙŠØªÙ… Ø§Ù„Ø¹Ø«ÙˆØ± Ø¹Ù„Ù‰ BOT_TOKEN ÙÙŠ Ù…ØªØºÙŠØ±Ø§Øª Ø§Ù„Ø¨ÙŠØ¦Ø©!")
        return
    
    application = Application.builder().token(BOT_TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU: [MessageHandler(filters.TEXT & ~filters.COMMAND, main_menu)],
            FIRST_NAME_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, first_name_input)],
            LAST_NAME_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, last_name_input)],
            VISIT_TYPE: [MessageHandler(filters.TEXT & ~filters.COMMAND, visit_type)],
            DOCTOR_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, doctor_name)],
            LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, location)],
            SPECIALTY: [MessageHandler(filters.TEXT & ~filters.COMMAND, specialty)],
            PRODUCTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, products)],
            COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, comment)],
            PHARMACY_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, pharmacy_name)],
            PHARMACY_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, pharmacy_address)],
            PHARMACY_PRODUCTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, pharmacy_products)],
            PHARMACY_COMMENT: [MessageHandler(filters.TEXT & ~filters.COMMAND, pharmacy_comment)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )
    
    application.add_handler(conv_handler)
    print("ğŸ¤– Ø§Ù„Ø¨ÙˆØª ÙŠØ¹Ù…Ù„ Ø§Ù„Ø¢Ù†...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)

if __name__ == "__main__":
    main()
