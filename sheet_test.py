from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

# =========================
# Create workbook and sheet
# =========================
wb = Workbook()
ws = wb.active
ws.title = "Daily Report_0"

# =========================
# Define styles
# =========================
header_fill = PatternFill("solid", fgColor="FFFF00")   # Yellow for headers
blue_fill = PatternFill("solid", fgColor="31859B")     # Blue for Name label
orange_fill = PatternFill("solid", fgColor="FABF8F")   # Orange for Date & separators
section_fill = PatternFill("solid", fgColor="C6E0B4")  # Greenish for A.M / P.M sections

center = Alignment(horizontal="center", vertical="center")  # Center text horizontally & vertically
bold = Font(bold=True)  # Bold font for headers/important cells

border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# =========================
# Title Section
# =========================
ws.merge_cells("A2:F2")        # Merge title across columns A to F
ws["A2"].value = "Daily Report"
ws["A2"].font = Font(bold=True, size=14)
ws["A2"].alignment = center
ws["A2"].fill = header_fill

# =========================
# Name & Date Section
# =========================
ws["A4"].value = "Name:"
ws["A5"].value = "Date:"
ws["A4"].font = ws["A5"].font = bold

# Merge cells for input fields next to labels
ws.merge_cells("B4:F4")
ws.merge_cells("B5:F5")

# Color the labels and input fields
for col in ["A", "B"]:
    ws[f"{col}4"].fill = blue_fill
    ws[f"{col}5"].fill = orange_fill

# =========================
# Table Header
# =========================
headers = ["A.M / P.M", "Doctor Name", "Hospital", "Specialist", "Product", "Comment"]
header_row = 7
for col, h in enumerate(headers, start=1):
    cell = ws.cell(row=header_row, column=col, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# =========================
# A.M Section
# =========================
ws.merge_cells("A8:A14")  # Merge A column for A.M
ws["A8"].value = "A.M"
ws["A8"].alignment = center
ws["A8"].font = bold
ws["A8"].fill = section_fill

# Add borders to A.M data cells (B-H)
for r in range(8, 15):
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = border

# Separator row after A.M
for c in range(1, 7):
    ws.cell(row=15, column=c).fill = orange_fill

# =========================
# P.M Section
# =========================
ws.merge_cells("A16:A28")  # Merge A column for P.M
ws["A16"].value = "P.M"
ws["A16"].alignment = center
ws["A16"].font = bold
ws["A16"].fill = section_fill

# Add borders to P.M data cells
for r in range(16, 29):
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = border

# Separator row after P.M
for c in range(1, 7):
    ws.cell(row=29, column=c).fill = orange_fill

# =========================
# Pharmacy Section
# =========================
ws.merge_cells("A30:A37")      # Merge column A for Pharmacy label
ws["A30"].value = "PHARMACY"
ws["A30"].alignment = center
ws["A30"].font = bold
ws["A30"].fill = header_fill

# Define pharmacy table headers
ph_headers = ["Pharmacy Name", "Address", "_" ,"Products", "Comments"]
for col, h in enumerate(ph_headers, start=2):
    if h in ["_", "Products"]:
        continue  # We'll merge these cells later
    cell = ws.cell(row=30, column=col, value=h)
    cell.font = bold
    cell.fill = header_fill
    cell.border = border
    cell.alignment = center

# Merge cells for Products column (D30:E37)
for r in range(30, 38):
    ws.merge_cells(f"D{r}:E{r}")

# Add the Products header
ws["D30"].value = "Products"
ws["D30"].font = bold
ws["D30"].fill = header_fill
ws["D30"].alignment = center
ws["D30"].border = border

# Add borders to pharmacy rows
for r in range(31, 38):
    for c in range(2, 7):
        ws.cell(row=r, column=c).border = border

# =========================
# Column Widths
# =========================
widths = [15, 25, 20, 20, 20, 30]  # Adjust column widths for better readability
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

# =========================
# Save the workbook
# =========================
wb.save("Daily_Report_00.xlsx")

# =========================

# ... (نفس الكود بتاعك)
ws["B5"].value = date.today().strftime("%d/%m/%Y") # بيضيف تاريخ اليوم تلقائياً
ws["B5"].alignment = Alignment(horizontal="left")
# =========================
